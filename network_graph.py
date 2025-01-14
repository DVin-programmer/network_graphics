#!/usr/bin/python3
# coding: utf8

'''
Программа для построения сетевых графиков
-----------------------------------------
Автор: Винокурова Д.В.
'''
from random import randint
from PySide2 import QtCore, QtGui, QtWidgets
from files.general_methods import StyleWidgets
import os, networkx as nx
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle, Font, Side, Border, Alignment
from datetime import datetime as dt
import webbrowser as wb
import platform
from winreg import OpenKey, QueryValueEx, HKEY_CURRENT_USER


# Преобразование LaTeX в http запрос
def http_text(strF):
    strF = strF.replace("\\", "%5C")
    strF = strF.replace("{", "%7B")
    strF = strF.replace("}", "%7D")
    strF = strF.replace(" ", "%20")
    strF = strF.replace("^", "%5E")
    strF = strF.replace("\n", "%0A")
    strF = strF.replace("&", "%26")

    strHTTP1 = "https://math.now.sh?from="
    strHTTP = strHTTP1 + strF + ".png"
    return strHTTP

# Генератор цвета
def color_generate():
    r = randint(43, 240)
    g = randint(72, 240)
    b = randint(70, 240)
    res_RGB = "#" + str(hex(r)[2:]) + str(hex(g)[2:]) + str(hex(b)[2:])
    return res_RGB

# Дата для Excel файла
def data_for_file():
    return dt.now().strftime("%Y-%m-%d_%H.%M.%S")

global flag_win_order_of_work
flag_win_order_of_work = False

class Ui_MainWindow(QtWidgets.QMainWindow):
    def __init__(self, *args, **kwargs):
        super(Ui_MainWindow, self).__init__(*args, **kwargs)
        #self.setFixedSize(750, 728)
        self.setupUi(self)
        self.setWindowTitle("Построение сетевого графика")

        # Флаг отвечает за сохранение таблицы в MS Excel
        self.flag_save_MSExcel = True
        # Флаг отвечает за загрузку данных из Excel
        self.flag_load_MSExcel = False


    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(572, 550)
        self.left = 30
        self.top = 70
        self.setGeometry(self.left, self.top, 572, 550)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.setMaximumSize(572, 550)
        MainWindow.setCentralWidget(self.centralwidget)

        # type_OS - тип установленной ОС. От этого зависит какого размера будет шрифт.
        type_OS = platform.system() + str(" ") + platform.release()
        type_OS = type_OS.lower()

        #type_OS = "Windows 7"
        #type_OS = type_OS.lower()
        if type_OS == "windows 7" or "windows 8" in type_OS:
            self.small_font = True
            #print("Умеьшенный шрифт")
        elif "windows 10" in type_OS:
            self.small_font = False
            #print("Нормальный шрифт")

        # Возвращаем бруезер, который устанолвен по умолчанию.
        # Браузеры Google Chrome, Mozilla Firefox корректно отображают информацию,
        # в остальных могут наблюдаться изменения.
        try:
            with OpenKey(HKEY_CURRENT_USER,
                         r"Software\\Microsoft\\Windows\\Shell\\Associations\\UrlAssociations\\http\\UserChoice") as key:
                browser = QueryValueEx(key, 'Progid')[0]
        except FileNotFoundError:
            self.font = QtGui.QFont()
            self.font.setFamily("Bahnschrift SemiLight SemiConde")
            self.font.setPointSize(16)

            self.label_browser = QtWidgets.QLabel(self.centralwidget)
            self.label_browser.setGeometry(QtCore.QRect(90, 165, 390, 50))
            self.label_browser.setObjectName("label")
            self.label_browser.setText("Для корректной работы установите по умолчанию \nGoogle Chrome или Mozilla Firefox!")
            self.label_browser.setStyleSheet("color: 'red'")
            self.label_browser.setFont(self.font)

        if not("Chrome" in browser or "Firefox" in browser):
            self.font = QtGui.QFont()
            self.font.setFamily("Bahnschrift SemiLight SemiConde")
            self.font.setPointSize(16)

            self.label_browser = QtWidgets.QLabel(self.centralwidget)
            self.label_browser.setGeometry(QtCore.QRect(90, 165, 390, 50))
            self.label_browser.setObjectName("label")
            self.label_browser.setText("Для корректной работы установите \nGoogle Chrome или Mozilla Firefox!")
            self.label_browser.setStyleSheet("color: 'red'")
            self.label_browser.setFont(self.font)

        # Устанавливаем размер шрифта в зависимости от типа ОС (флаг self.small_font)
        if self.small_font == False:
            self.font = QtGui.QFont()
            self.font.setFamily("Bahnschrift SemiLight SemiConde")
            self.font.setPointSize(12)
        else:
            self.font = QtGui.QFont()
            self.font.setFamily("Bahnschrift SemiLight SemiConde")
            self.font.setPointSize(10)


        self.setWindowIcon(QtGui.QIcon("files\\ico\\NetDiag.ico"))

        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(20, 38, 181, 20))
        self.label.setObjectName("label")
        self.label.setText("Выберите количество работ:")
        self.label.setFont(self.font)

        self.spinBox_num_jobs = QtWidgets.QSpinBox(self.centralwidget)
        self.spinBox_num_jobs.setGeometry(QtCore.QRect(210, 39, 42, 22))
        self.spinBox_num_jobs.setValue(5)
        self.spinBox_num_jobs.setObjectName("spinBox_num_jobs")
        self.spinBox_num_jobs.setFont(self.font)

        self.pushButton_next = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_next.setGeometry(QtCore.QRect(270, 36, 75, 27))
        self.pushButton_next.setObjectName("pushButton_next")
        self.pushButton_next.clicked.connect(self.set_flag)
        self.pushButton_next.setText("Далее")
        self.pushButton_next.setFont(self.font)
        background_color = "#51d77a"
        border_color = "#2d8849"
        color = "#2d8849"
        background_color_hover = "#75e998"
        border_color_hover = "#2d8849"
        color_hover = "#2d8849"
        background_color_press = "#17c14b"
        border_color_press = "#2d8849"
        color_press = "#75e998"
        border_width = "1px"
        border_radius = "2px"
        button_style = StyleWidgets()
        button_style.properties_button(self.pushButton_next, background_color,
                                    border_color, color,
                                    background_color_hover, border_color_hover, color_hover,
                                    background_color_press, border_color_press, color_press,
                                    border_width, border_radius)

        self.pushButton_reset = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_reset.setGeometry(QtCore.QRect(370, 36, 180, 27))
        self.pushButton_reset.setObjectName("pushButton_reset")
        self.pushButton_reset.setText("Создать новую таблицу")
        self.pushButton_reset.setFont(self.font)
        self.pushButton_reset.clicked.connect(self.clear_info)
        background_color = "#fa7070"
        border_color = "#9a2222"
        color = "#9a2222"
        background_color_hover = "#f3a5a5"
        border_color_hover = "#9a2222"
        color_hover = "#9a2222"
        background_color_press = "#d33a3a"
        border_color_press = "#9a2222"
        color_press = "#f3a5a5"
        border_width = "1px"
        border_radius = "2px"
        button_style.properties_button(self.pushButton_reset, background_color,
                                    border_color, color,
                                    background_color_hover, border_color_hover, color_hover,
                                    background_color_press, border_color_press, color_press,
                                    border_width, border_radius)

        self.table_jobs = QtWidgets.QTableWidget(self.centralwidget)  # Создаём таблицу
        self.table_jobs.setFont(self.font)
        self.table_jobs.setColumnCount(3)  # Устанавливаем 5 столбцов
        # Прописываем наименование столбцов
        self.table_jobs.setHorizontalHeaderLabels(["Номер \n работы", "Описание работы",
                                                        "Длительность \n(в днях)"])
        self.table_jobs.setVisible(False)
        self.table_jobs.resizeColumnsToContents()
        self.table_jobs.resizeRowsToContents()
        self.table_jobs.setGeometry(QtCore.QRect(20, 120, 530, 285))
        self.table_jobs.setColumnWidth(1, 340)


        self.pushButton_order_of_work = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_order_of_work.setText("Установить порядок \n выполнения работ")
        self.pushButton_order_of_work.setGeometry(QtCore.QRect(380, 420, 150, 54))
        self.pushButton_order_of_work.setObjectName("pushButton_order_of_work")
        self.pushButton_order_of_work.clicked.connect(self.check_table)
        self.pushButton_order_of_work.setFont(self.font)
        background_color = "#6cbef1"
        border_color = "#155b87"
        color = "#155b87"
        background_color_hover = "#bbe1f9"
        border_color_hover = "#155b87"
        color_hover = "#155b87"
        background_color_press = "#3a84b3"
        border_color_press = "#3a84b3"
        color_press = "#155b87"
        border_width = "1px"
        border_radius = "2px"
        button_style = StyleWidgets()
        button_style.properties_button(self.pushButton_order_of_work, background_color,
                                       border_color, color,
                                       background_color_hover, border_color_hover, color_hover,
                                       background_color_press, border_color_press, color_press,
                                       border_width, border_radius)
        self.pushButton_order_of_work.setVisible(False)


        if self.small_font == False:
            self.font = QtGui.QFont()
            self.font.setFamily("Bahnschrift SemiLight SemiConde")
            self.font.setPointSize(11)
        else:
            self.font = QtGui.QFont()
            self.font.setFamily("Bahnschrift SemiLight SemiConde")
            self.font.setPointSize(9)
        self.label_author = QtWidgets.QLabel(self.centralwidget)
        self.label_author.setGeometry(QtCore.QRect(20, 490, 181, 20))
        self.label_author.setObjectName("label_author")
        self.label_author.setText("© Винокурова Д.В., 2020")
        self.label_author.setFont(self.font)

        self.order_of_work = Win_OrderOfWork(self)

        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 572, 21))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)

        self.action_instruction = QtWidgets.QAction(MainWindow)
        self.action_instruction.triggered.connect(self.open_instruction)
        self.action_instruction.setText("Инструкция")
        self.menubar.addAction(self.action_instruction)

        self.action_save_MSExcel = QtWidgets.QAction(MainWindow)
        self.action_save_MSExcel.triggered.connect(self.save_MSExcel)
        self.action_save_MSExcel.setText("Сохранить таблицу в Excel ")
        self.menubar.addAction(self.action_save_MSExcel)

        self.action_download_MSExcel = QtWidgets.QAction(MainWindow)
        self.action_download_MSExcel.triggered.connect(self.download_MSExcel)
        self.action_download_MSExcel.setText("Загрузить таблицу из файла")
        self.menubar.addAction(self.action_download_MSExcel)


        bar = self.menuBar()
        self.delete = bar.addMenu("Удалить ...")

        self.action_delete_file_MSExcel = QtWidgets.QAction(MainWindow)
        self.action_delete_file_MSExcel.triggered.connect(self.delete_file_MSExcel)
        self.action_delete_file_MSExcel.setText("Удалить файл")
        self.delete.addAction(self.action_delete_file_MSExcel)
        #self.menubar.addAction(self.action_delete_file_MSExcel)

        self.action_delete_files = QtWidgets.QAction(MainWindow)
        self.action_delete_files.triggered.connect(self.delete_files)
        self.action_delete_files.setText("Удалить все файлы")
        #self.menubar.addAction(self.action_delete_files)
        self.delete.addAction(self.action_delete_files)


        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        self.delete_excel_file = Win_DeleteExcelFile(self)


    def closeEvent(self, event):
        '''
        Запрос разрешения на закрытие программы.
        '''
        self.main_win_dialog_yes_no = QtWidgets.QMessageBox()
        self.main_win_dialog_yes_no.addButton("Да", QtWidgets.QMessageBox.AcceptRole)
        self.main_win_dialog_yes_no.addButton("Нет", QtWidgets.QMessageBox.AcceptRole)
        self.main_win_dialog_yes_no.setIcon(QtWidgets.QMessageBox.Information)
        self.main_win_dialog_yes_no.setWindowTitle("Удаление")
        self.main_win_dialog_yes_no.setWindowIcon(QtGui.QIcon("files\\ico\\NetDiag.ico"))

        self.left = 150
        self.top = 210
        self.main_win_dialog_yes_no.setGeometry(self.left, self.top, 100, 550)
        self.main_win_dialog_yes_no.setText("Вы уверены, что хотите закрыть программу?")
        self.main_win_dialog_yes_no.setFont(self.font)
        self.main_win_dialog_yes_no.exec()

        if self.main_win_dialog_yes_no.clickedButton().text() == "Да":
            event.accept()
        elif self.main_win_dialog_yes_no.clickedButton().text() == "Нет":
            event.ignore()


    def open_instruction(self):
        '''
        Загружается инструкция в *.pdf
        '''
        print("открывается инструкция")
        wb.open("files\\7_SegmentIndicator.pdf", new=2)


    def save_MSExcel(self):
        '''
        Метод позволяет сохранять таблицу в MS Excel.
        '''
        self.delete_excel_file.close()
        try:
            self.msg_box_save_MSExcel = QtWidgets.QMessageBox()
            self.msg_box_save_MSExcel.setFixedSize(100, 500)
            self.msg_box_save_MSExcel.adjustSize()
            self.msg_box_save_MSExcel.setIcon(QtWidgets.QMessageBox.Warning)
            self.msg_box_save_MSExcel.setWindowIcon(QtGui.QIcon("files\\ico\\NetDiag.ico"))
            self.msg_box_save_MSExcel.setWindowTitle("Предупреждение")
            if self.flag_save_MSExcel == True:
                wb = Workbook()
                sheet = wb.active
                sheet.title = "Проводимые работы"

                # Задаем стиль1
                ns1 = NamedStyle(name="style1")
                ns1.font = Font(bold=True, size=12)
                border = Side(style="thin", color="000000")
                ns1.border = Border(left=border, top=border,
                                    right=border, bottom=border)
                ns1.alignment = Alignment(wrap_text=True, horizontal="center",
                                          vertical="center")
                wb.add_named_style(ns1)
                # ---------------------
                # Задаем стиль2
                ns2 = NamedStyle(name="style2")
                ns2.font = Font(size=12)
                border = Side(style="thin", color="000000")
                ns2.border = Border(left=border, top=border,
                                    right=border, bottom=border)
                ns2.alignment = Alignment(horizontal="center",
                                          vertical="center")
                wb.add_named_style(ns2)
                # ---------------------
                # Задаем стиль2
                ns3 = NamedStyle(name="style3")
                ns3.font = Font(size=12)
                border = Side(style="thin", color="000000")
                ns3.border = Border(left=border, top=border,
                                    right=border, bottom=border)
                ns3.alignment = Alignment(wrap_text=True, horizontal="left",
                                          vertical="center")
                wb.add_named_style(ns3)

                sheet.column_dimensions["B"].width = 42
                sheet.column_dimensions["C"].width = 15
                sheet.row_dimensions[1].height = 30

                row = 1
                sheet["A" + str(row)].style = "style1"
                sheet["B" + str(row)].style = "style1"
                sheet["C" + str(row)].style = "style1"

                sheet["A" + str(row)] = "Номер\nработы"
                sheet["B" + str(row)] = "Описание работы"
                sheet["C" + str(row)] = "Длительность\n(дней)"

                for i in range(0, self.num_str_in_table):
                    row += 1
                    sheet["A" + str(row)].style = "style2"
                    sheet["B" + str(row)].style = "style3"
                    sheet["C" + str(row)].style = "style2"

                    sheet["A" + str(row)] = str(i + 1)
                    sheet["B" + str(row)] = lst_work_description[i]
                    sheet["C" + str(row)] = lst_work_day[i]


                entered_file_name, ok = QtWidgets.QInputDialog.getText(self, 'Имя файла',
                                                                       'Введите имя файла:')
                if ok and entered_file_name != "":
                    fileName = "files\\excel_files\\" + str(entered_file_name) + ".xlsx"
                    wb.save(fileName)

                    self.left = 160
                    self.top = 270
                    self.msg_box_save_MSExcel.setIcon(QtWidgets.QMessageBox.Information)
                    self.msg_box_save_MSExcel.setWindowTitle("Информация")
                    self.msg_box_save_MSExcel.setWindowIcon(QtGui.QIcon("files\\ico\\NetDiag.ico"))
                    self.msg_box_save_MSExcel.setGeometry(self.left, self.top, 572, 550)
                    self.msg_box_save_MSExcel.setText("Файл «" + str(entered_file_name) +\
                                                      ".xlsx» \nуспешно сохранен!")
                    self.msg_box_save_MSExcel.setFont(self.font)
                    self.msg_box_save_MSExcel.exec()
                elif entered_file_name == "":
                    self.left = 140
                    self.top = 270
                    self.msg_box_save_MSExcel.setIcon(QtWidgets.QMessageBox.Information)
                    self.msg_box_save_MSExcel.setWindowTitle("Информация")
                    self.msg_box_save_MSExcel.setWindowIcon(QtGui.QIcon("files\\ico\\NetDiag.ico"))
                    self.msg_box_save_MSExcel.setGeometry(self.left, self.top, 572, 550)
                    self.msg_box_save_MSExcel.setText("Имя файла не может быть пустым!")
                    self.msg_box_save_MSExcel.setFont(self.font)
                    self.msg_box_save_MSExcel.exec()
                else:
                    self.left = 140
                    self.top = 270
                    self.msg_box_save_MSExcel.setIcon(QtWidgets.QMessageBox.Information)
                    self.msg_box_save_MSExcel.setWindowTitle("Информация")
                    self.msg_box_save_MSExcel.setWindowIcon(QtGui.QIcon("files\\ico\\NetDiag.ico"))
                    self.msg_box_save_MSExcel.setGeometry(self.left, self.top, 572, 550)
                    self.msg_box_save_MSExcel.setText("Вы не ввели имя файла. Файл не сохранен!")
                    self.msg_box_save_MSExcel.setFont(self.font)
                    self.msg_box_save_MSExcel.exec()
            else:
                self.left = 80
                self.top = 270
                self.msg_box_save_MSExcel.setIcon(QtWidgets.QMessageBox.Warning)
                self.msg_box_save_MSExcel.setWindowTitle("Предупреждение")
                self.msg_box_save_MSExcel.setGeometry(self.left, self.top, 572, 550)
                self.msg_box_save_MSExcel.setText("Создайте таблицу и нажмите кнопку " +\
                                         "«Установить порядок \nвыполнения работ».")
                self.msg_box_save_MSExcel.setFont(self.font)
                self.msg_box_save_MSExcel.exec()

        except (KeyError, NameError):
            self.left = 80
            self.top = 270
            self.msg_box_save_MSExcel.setIcon(QtWidgets.QMessageBox.Warning)
            self.msg_box_save_MSExcel.setWindowTitle("Предупреждение")
            self.msg_box_save_MSExcel.setGeometry(self.left, self.top, 572, 550)
            self.msg_box_save_MSExcel.setText("Таблица не проверена! Нажмите кнопку " +\
                                         "«Установить порядок \nвыполнения работ».")
            self.msg_box_save_MSExcel.setFont(self.font)
            self.msg_box_save_MSExcel.exec()
        except PermissionError:
            self.left = 230
            self.top = 270
            self.msg_box_save_MSExcel.setIcon(QtWidgets.QMessageBox.Warning)
            self.msg_box_save_MSExcel.setWindowTitle("Предупреждение")
            self.msg_box_save_MSExcel.setGeometry(self.left, self.top, 572, 550)
            self.msg_box_save_MSExcel.setText("Закройте MS Excel!")
            self.msg_box_save_MSExcel.setFont(self.font)
            self.msg_box_save_MSExcel.exec()
        except IndexError:
            self.left = 80
            self.top = 270
            self.msg_box_save_MSExcel.setIcon(QtWidgets.QMessageBox.Warning)
            self.msg_box_save_MSExcel.setWindowTitle("Предупреждение")
            self.msg_box_save_MSExcel.setGeometry(self.left, self.top, 572, 550)
            self.msg_box_save_MSExcel.setText("Заполните таблицу и нажмите на кнопку " +\
                                         "«Установить порядок \nвыполнения работ».")
            self.msg_box_save_MSExcel.setFont(self.font)
            self.msg_box_save_MSExcel.exec()
        except AttributeError:
            self.left = 230
            self.top = 270
            self.msg_box_save_MSExcel.setIcon(QtWidgets.QMessageBox.Warning)
            self.msg_box_save_MSExcel.setWindowTitle("Предупреждение")
            self.msg_box_save_MSExcel.setGeometry(self.left, self.top, 572, 550)
            self.msg_box_save_MSExcel.setText("Создайте таблицу!")
            self.msg_box_save_MSExcel.setFont(self.font)
            self.msg_box_save_MSExcel.exec()


    def download_MSExcel(self):
        '''
        Метод позволяет загружать таблицу из MS Excel.
        '''
        self.delete_excel_file.close()

        self.clear_info()

        self.msg_box_load_MSExcel = QtWidgets.QMessageBox()
        self.msg_box_load_MSExcel.setFixedSize(100, 500)
        self.msg_box_load_MSExcel.adjustSize()
        self.msg_box_load_MSExcel.setIcon(QtWidgets.QMessageBox.Warning)
        self.msg_box_load_MSExcel.setWindowIcon(QtGui.QIcon("files\\ico\\NetDiag.ico"))
        self.msg_box_load_MSExcel.setWindowTitle("Предупреждение")

        file_name_exel = QtWidgets.QFileDialog.getOpenFileName(self, 'Загрузить файл',
                                                               'files\\excel_files',
                                                               "*.xlsx")[0]
        if file_name_exel != "":
            wb = load_workbook(file_name_exel)
            sheet = wb.active

            global lst_MSExcel_A, lst_MSExcel_B, lst_MSExcel_C
            lst_MSExcel_A = []
            lst_MSExcel_B = []
            lst_MSExcel_C = []
            i = 1
            while True:
                col_A = sheet['A' + str(i)].value
                col_B = sheet['B' + str(i)].value
                col_C = sheet['C' + str(i)].value
                if col_A == None:
                    break
                lst_MSExcel_A.append(col_A)
                lst_MSExcel_B.append(str(col_B))
                lst_MSExcel_C.append(col_C)
                i += 1
            if i <= 4:
                self.left = 70
                self.top = 270
                self.msg_box_load_MSExcel.setGeometry(self.left, self.top, 572, 550)
                self.msg_box_load_MSExcel.setText("Неверно заполнен файл «" + file_name_exel + "»!")
                self.msg_box_load_MSExcel.setFont(self.font)
                self.msg_box_load_MSExcel.exec()
            else:
                self.num_str_in_table = i - 2
                self.spinBox_num_jobs.setValue(i - 2)
                self.flag_load_MSExcel = True
                self.create_table()

            '''print(lst_MSExcel_A)
            print(lst_MSExcel_B)
            print(lst_MSExcel_C)'''


    def delete_file_MSExcel(self):
        '''
        Метод запускает метод удаления excel
        файла из класса Win_DeleteExcelFile
        '''
        self.delete_excel_file.update_combobox()
        self.delete_excel_file.show()


    def delete_files(self):
        '''
        Удаление всех Excel файлов в папке excel_files
        '''
        self.delete_excel_file.close()

        self.delete_files_yes_no = QtWidgets.QMessageBox()
        self.delete_files_yes_no.addButton("Да", QtWidgets.QMessageBox.AcceptRole)
        self.delete_files_yes_no.addButton("Нет", QtWidgets.QMessageBox.AcceptRole)
        self.delete_files_yes_no.setIcon(QtWidgets.QMessageBox.Information)
        self.delete_files_yes_no.setWindowTitle("Удаление")
        self.delete_files_yes_no.setWindowIcon(QtGui.QIcon("files\\ico\\NetDiag.ico"))

        self.msg_box_delete_files_info = QtWidgets.QMessageBox()
        self.msg_box_delete_files_info.setIcon(QtWidgets.QMessageBox.Information)
        self.msg_box_delete_files_info.setWindowTitle("Информация")
        self.msg_box_delete_files_info.setWindowIcon(QtGui.QIcon("files\\ico\\sucessfull.ico"))

        self.lst_files_for_delete = os.listdir("files\\excel_files")
        if len(self.lst_files_for_delete) == 0:
            self.left = 180
            self.top = 210
            self.msg_box_delete_files_info.setGeometry(self.left, self.top, 100, 550)
            self.msg_box_delete_files_info.setText("Файлы отсутсвуют!")
            self.msg_box_delete_files_info.setWindowIcon(QtGui.QIcon("files\\ico\\NetDiag.ico"))
            self.msg_box_delete_files_info.setFont(self.font)
            self.msg_box_delete_files_info.exec()
        else:
            self.left = 160
            self.top = 210
            self.delete_files_yes_no.setGeometry(self.left, self.top, 100, 550)
            self.delete_files_yes_no.setText("Вы уверены, что хотите удалить все файлы?\n"
                                             "Это действие нельзя будет отменить.")
            self.delete_files_yes_no.setFont(self.font)
            self.delete_files_yes_no.exec()

            if self.delete_files_yes_no.clickedButton().text() == "Да":
                self.res_delete_files_yes_no = True
            elif self.delete_files_yes_no.clickedButton().text() == "Нет":
                self.res_delete_files_yes_no = False

            if self.res_delete_files_yes_no:
                for i in range(0, len(self.lst_files_for_delete)):
                    os.remove("files\\excel_files\\" + str(self.lst_files_for_delete[i]))

                self.left = 150
                self.top = 210
                self.msg_box_delete_files_info.setGeometry(self.left, self.top, 100, 550)
                self.msg_box_delete_files_info.setText("Файлы успешно удалены!")
                self.msg_box_delete_files_info.setFont(self.font)
                self.msg_box_delete_files_info.exec()




    def set_flag(self):
        '''
        Метод необходим, чтобы установить флаг, который говорит о том, что
        данные вводятся вручную.
        '''
        self.flag_load_MSExcel = False
        self.create_table()


    def create_table(self):
        '''
        Создание таблицы с необходимым количеством строк,
        указанных в spinBox
        '''
        '''self.font = QtGui.QFont()
        self.font.setFamily("Bahnschrift SemiLight SemiConde")
        self.font.setPointSize(11)'''
        if self.small_font == False:
            self.font = QtGui.QFont()
            self.font.setFamily("Bahnschrift SemiLight SemiConde")
            self.font.setPointSize(11)
        else:
            self.font = QtGui.QFont()
            self.font.setFamily("Bahnschrift SemiLight SemiConde")
            self.font.setPointSize(10)

        self.msg_box = QtWidgets.QMessageBox()
        self.msg_box.setFixedSize(100, 500)
        self.msg_box.adjustSize()
        self.msg_box.setWindowIcon(QtGui.QIcon("files\\ico\\NetDiag.ico"))

        self.num_str_in_table = self.spinBox_num_jobs.value()
        if 0 <= self.num_str_in_table <= 4:
            print("Введите большее количество работ")
            self.msg_box.setIcon(QtWidgets.QMessageBox.Warning)
            self.msg_box.setWindowTitle("Предупреждение")
            self.msg_box.setText("Введите большее количество работ!")
            self.msg_box.setFont(self.font)
            self.msg_box.exec()
        else:
            # Делаем кнопку неактивной, задаем стиль
            self.pushButton_next.setEnabled(False)
            self.pushButton_next.setStyleSheet('''
                    background-color: "#cccccc";
                    border-style: solid;
                    border-color: "#909090";
                    border-width: 1px;
                    border-radius: 2px;
                    color: "#909090";''')

            self.spinBox_num_jobs.setEnabled(False)
            self.pushButton_order_of_work.setVisible(True)

            self.table_jobs.setVisible(True)
            self.table_jobs.setRowCount(self.num_str_in_table)  # строки
            self.table_jobs.setFont(self.font)

            # Удаление номеров
            self.table_jobs.verticalHeader().setVisible(False)
            self.table_jobs.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
            # Устанавливаем выравнивание на заголовки
            for i in range(0, 3):
                self.table_jobs.horizontalHeaderItem(i).setTextAlignment(QtCore.Qt.AlignHCenter)

                # Заполнение номеров работ
                for i in range(0, self.num_str_in_table):
                    self.num_jobs_table = QtWidgets.QTableWidgetItem(str(i + 1))
                    # print(i)
                    self.table_jobs.setItem(i, 0, self.num_jobs_table)
                    self.num_jobs_table.setTextAlignment(QtCore.Qt.AlignCenter)
                    self.num_jobs_table.setFlags(QtCore.Qt.ItemIsEnabled)

            if self.flag_load_MSExcel == False:
                #print("IF")
                # Заполнение остальных ячеек
                for i in range(0, self.num_str_in_table):
                    for j in range(1, 3):
                        self.cell = QtWidgets.QTableWidgetItem("")
                        self.table_jobs.setItem(i, j, self.cell)
                        self.cell.setTextAlignment(QtCore.Qt.AlignCenter)
                        # cell.setFlags(QtCore.Qt.ItemIsEnabled)
            else:
                #print("ELSE")
                # Заполнение столбца "Описание" из Excel
                for i in range(0, self.num_str_in_table):
                    self.coll_B = QtWidgets.QTableWidgetItem(lst_MSExcel_B[i + 1])
                    self.table_jobs.setItem(i, 1, self.coll_B)
                    self.coll_B.setTextAlignment(QtCore.Qt.AlignCenter)

                # Заполнение столбца "Длительность" из Excel
                for i in range(0, self.num_str_in_table):
                    self.coll_C = QtWidgets.QTableWidgetItem(str(lst_MSExcel_C[i + 1]))
                    self.table_jobs.setItem(i, 2, self.coll_C)
                    self.coll_C.setTextAlignment(QtCore.Qt.AlignCenter)


    def check_table(self):
        '''
        Проверяет поля таблицы на корректность.
        '''
        # Проверка столбца "Описание работы"
        flag_col2 = False
        flag_col3 = False
        global lst_work_description
        lst_work_description = []
        for i in range(0, self.num_str_in_table):
            cell = self.table_jobs.item(i, 1).text()
            if cell == "":
                self.table_jobs.item(i, 1).setBackground(QtGui.QColor("#f78989"))
                flag_col2 = True
            else:
                self.table_jobs.item(i, 1).setBackground(QtGui.QColor("#ffffff"))
                lst_work_description.append(cell)

        # Проверка столбца "Длительность"
        global lst_work_day
        lst_work_day = []
        self.table_jobs.setStyleSheet("QTableWidget::item:selected{ background-color: #4f7cfc}")
        for i in range(0, self.num_str_in_table):
            cell = self.table_jobs.item(i, 2).text()
            if cell == "" or not(cell.isdigit()):
                self.table_jobs.item(i, 2).setBackground(QtGui.QColor("#f78989"))
                flag_col3 = True
            else:
                self.table_jobs.item(i, 2).setBackground(QtGui.QColor("#ffffff"))
                lst_work_day.append(int(cell))

        if flag_col2 == False and flag_col3 == False:
            self.open_win_order_of_work()
            self.flag_save_MSExcel = True
            '''for i in range(0, self.num_str_in_table):
                for j in range(1, 3):
                    cell.setFlags(QtCore.Qt.ItemIsEnabled)'''
        else:
            self.left = 150
            self.top = 270
            self.msg_box.setWindowTitle("Предупреждение")
            self.msg_box.setGeometry(self.left, self.top, 572, 550)
            self.msg_box.setText("Необходимо заполнить и проверить данные\n" +\
                                 "в ячейках, которые отмечены красным цветом!")
            self.msg_box.setFont(self.font)
            self.msg_box.exec()

    def clear_info(self):
        '''
        Метод возвращает окно в исходное состояние
        '''
        #print("clear_info:",self.flag_load_MSExcel)
        if self.flag_load_MSExcel == True:
            for i in range(0, self.num_str_in_table):
                for j in range(1, 3):
                    self.cell = QtWidgets.QTableWidgetItem("")
                    self.table_jobs.setItem(i, j, self.cell)
                    self.cell.setTextAlignment(QtCore.Qt.AlignCenter)

        self.flag_save_MSExcel = False
        self.flag_load_MSExcel = False


        self.table_jobs.setRowCount(0)
        self.table_jobs.setVisible(False)
        self.pushButton_order_of_work.setVisible(False)
        self.spinBox_num_jobs.setEnabled(True)

        self.pushButton_order_of_work.setEnabled(True)
        background_color = "#6cbef1"
        border_color = "#155b87"
        color = "#155b87"
        background_color_hover = "#bbe1f9"
        border_color_hover = "#155b87"
        color_hover = "#155b87"
        background_color_press = "#3a84b3"
        border_color_press = "#3a84b3"
        color_press = "#155b87"
        border_width = "1px"
        border_radius = "2px"
        button_style = StyleWidgets()
        button_style.properties_button(self.pushButton_order_of_work, background_color,
                                       border_color, color,
                                       background_color_hover, border_color_hover, color_hover,
                                       background_color_press, border_color_press, color_press,
                                       border_width, border_radius)
        self.pushButton_order_of_work.setVisible(False)


        self.pushButton_next.setEnabled(True)
        background_color = "#51d77a"
        border_color = "#2d8849"
        color = "#2d8849"
        background_color_hover = "#75e998"
        border_color_hover = "#2d8849"
        color_hover = "#2d8849"
        background_color_press = "#17c14b"
        border_color_press = "#2d8849"
        color_press = "#75e998"
        border_width = "1px"
        border_radius = "2px"
        button_style = StyleWidgets()
        button_style.properties_button(self.pushButton_next, background_color,
                                       border_color, color,
                                       background_color_hover, border_color_hover, color_hover,
                                       background_color_press, border_color_press, color_press,
                                       border_width, border_radius)

        self.order_of_work.close()

    def open_win_order_of_work(self):
        self.MSExel = True

        if self.small_font == False:
            self.font = QtGui.QFont()
            self.font.setFamily("Bahnschrift SemiLight SemiConde")
            self.font.setPointSize(11)
        else:
            self.font = QtGui.QFont()
            self.font.setFamily("Bahnschrift SemiLight SemiConde")
            self.font.setPointSize(10)

        self.table_jobs.horizontalHeaderItem(0).setFont(self.font)
        self.table_jobs.horizontalHeaderItem(1).setFont(self.font)
        self.table_jobs.horizontalHeaderItem(2).setFont(self.font)

        self.order_of_work.show()
        self.order_of_work.add_buttons(self.num_str_in_table, self.small_font)


# Класс окна для удаления Excel файлов
class Win_DeleteExcelFile(QtWidgets.QMainWindow):
    def __init__(self, window = None):
        self.window = window
        super(Win_DeleteExcelFile, self).__init__(window)
        self.show_files(self)

    def show_files(self, MainWindow):
        '''
        Метод отображает окно с combobox для выбора
        необходимого файла для удаления
        '''
        if self.window.small_font == False:
            self.font = QtGui.QFont()
            self.font.setFamily("Bahnschrift SemiLight SemiConde")
            self.font.setPointSize(12)
        else:
            self.font = QtGui.QFont()
            self.font.setFamily("Bahnschrift SemiLight SemiConde")
            self.font.setPointSize(10)

        self.setWindowTitle("Удаление файлов")
        # self.resize(700, 550)
        self.left = 610
        self.top = 170
        self.setGeometry(self.left, self.top, 400, 200)
        self.setWindowIcon(QtGui.QIcon("files\\ico\\delete_excel.ico"))
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        MainWindow.setCentralWidget(self.centralwidget)

        self.label_file_selection = QtWidgets.QLabel(self.centralWidget())
        self.label_file_selection.setFont(self.font)
        self.label_file_selection.setGeometry(QtCore.QRect(60, 20, 390, 50))
        self.label_file_selection.setObjectName("label_file_selection")
        self.label_file_selection.setText("Выберите один из файлов:")


        self.lst_all_files = os.listdir("files\\excel_files")
        self.combobox_files = QtWidgets.QComboBox(self.centralwidget)
        self.combobox_files.setGeometry(QtCore.QRect(60, 70, 250, 25))
        self.combobox_files.setFont(self.font)
        self.selected_file_combobox = self.combobox_files.currentText()
        self.combobox_files.addItems(self.lst_all_files)
        self.combobox_files.activated.connect(self.selected_file)

        self.pushButton_delete_file = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_delete_file.setGeometry(QtCore.QRect(230, 120, 80, 27))
        self.pushButton_delete_file.setObjectName("pushButton_reset")
        self.pushButton_delete_file.setFont(self.font)
        self.pushButton_delete_file.setText("Удалить")
        self.pushButton_delete_file.clicked.connect(self.delete_file)
        background_color = "#fa7070"
        border_color = "#9a2222"
        color = "#9a2222"
        background_color_hover = "#f3a5a5"
        border_color_hover = "#9a2222"
        color_hover = "#9a2222"
        background_color_press = "#d33a3a"
        border_color_press = "#9a2222"
        color_press = "#f3a5a5"
        border_width = "1px"
        border_radius = "2px"
        button_style = StyleWidgets()
        button_style.properties_button(self.pushButton_delete_file, background_color,
                                       border_color, color,
                                       background_color_hover, border_color_hover, color_hover,
                                       background_color_press, border_color_press, color_press,
                                       border_width, border_radius)

        self.msg_box_delete_file = QtWidgets.QMessageBox()
        self.msg_box_delete_file.setFixedSize(100, 500)
        self.msg_box_delete_file.adjustSize()
        self.msg_box_delete_file.setIcon(QtWidgets.QMessageBox.Warning)
        self.msg_box_delete_file.setWindowIcon(QtGui.QIcon("files\\ico\\delete_excel.ico"))
        self.msg_box_delete_file.setWindowTitle("Предупреждение")

        self.msg_box_delete_file_info = QtWidgets.QMessageBox()
        self.msg_box_delete_file_info.setIcon(QtWidgets.QMessageBox.Information)
        self.msg_box_delete_file_info.setWindowTitle("Информация")
        self.msg_box_delete_file_info.setWindowIcon(QtGui.QIcon("files\\ico\\sucessfull.ico"))

        self.dialog_yes_no = QtWidgets.QMessageBox()
        self.dialog_yes_no.addButton("Да", QtWidgets.QMessageBox.AcceptRole)
        self.dialog_yes_no.addButton("Нет", QtWidgets.QMessageBox.AcceptRole)
        self.dialog_yes_no.setIcon(QtWidgets.QMessageBox.Information)
        self.dialog_yes_no.setWindowTitle("Удаление")
        self.dialog_yes_no.setWindowIcon(QtGui.QIcon("files\\ico\\delete_excel.ico"))


    def selected_file(self):
        '''
        Выбор названия файла из списка combobox.
        '''
        self.selected_file_combobox = self.combobox_files.currentText()

    def delete_file(self):
        '''
        Удаление excel файла.
        '''
        if self.selected_file_combobox == "":
            self.left = 710
            self.top = 210
            self.msg_box_delete_file.setGeometry(self.left, self.top, 572, 550)
            self.msg_box_delete_file.setText("Выберите файл!")
            self.msg_box_delete_file.setFont(self.font)
            self.msg_box_delete_file.exec()
        else:
            self.left = 550
            self.top = 210
            self.dialog_yes_no.setGeometry(self.left, self.top, 100, 550)
            self.dialog_yes_no.setText("Вы уверены, что хотите удалить файл «" +\
                                                   self.selected_file_combobox + "»?")
            self.dialog_yes_no.setFont(self.font)
            self.dialog_yes_no.exec()

            if self.dialog_yes_no.clickedButton().text() == "Да":
                self.res_dialog_yes_no = True
            elif self.dialog_yes_no.clickedButton().text() == "Нет":
                self.res_dialog_yes_no = False

            if self.res_dialog_yes_no:
                os.remove("files\\excel_files\\" + str(self.selected_file_combobox))
                self.update_combobox()

                self.left = 630
                self.top = 210
                self.msg_box_delete_file_info.setGeometry(self.left, self.top, 100, 550)
                self.msg_box_delete_file_info.setText("Файл «" + self.selected_file_combobox +\
                                                      "» \nуспешно удален!")
                self.msg_box_delete_file_info.setFont(self.font)
                self.msg_box_delete_file_info.exec()

                self.selected_file_combobox = self.combobox_files.currentText()


    def update_combobox(self):
        self.combobox_files.clear()
        self.lst_all_files = os.listdir("files\\excel_files")
        self.lst_all_files.insert(0, "")
        self.combobox_files.addItems(self.lst_all_files)



# Класс для окна с порядком выполнения работ
class Win_OrderOfWork(QtWidgets.QMainWindow):
    def __init__(self, window = None):
        #QtWidgets.QWidget.__init__(self)
        super(Win_OrderOfWork, self).__init__(window)
        self.window = window
        self.msg_box_win_order_of_work = QtWidgets.QMessageBox()
        self.msg_box_win_order_of_work.setFixedSize(100, 500)
        self.msg_box_win_order_of_work.adjustSize()

        # Пока дочернее окно открыто запрещаем переход на главное
        # self.setWindowModality(QtCore.Qt.ApplicationModal)

        self.building = BuildingNetworkGraph()

        # Запрет нажатия крестика
        # self.setWindowFlag(QtCore.Qt.WindowCloseButtonHint, False)


    def add_buttons(self, num_str_in_table, small_font):
        '''
        Данный метод размещает в окне два виджета и
        размещает кнопки для выбора работ.
        '''
        if small_font == False:
            self.font = QtGui.QFont()
            self.font.setFamily("Bahnschrift SemiLight SemiConde")
            self.font.setPointSize(12)
        else:
            self.font = QtGui.QFont()
            self.font.setFamily("Bahnschrift SemiLight SemiConde")
            self.font.setPointSize(10)

        self.setWindowTitle("Порядок выполнения работ")
        #self.resize(700, 550)
        self.left = 610
        self.top = 70
        self.setGeometry(self.left, self.top, 700, 550)

        #self.setMaximumSize(1000, 700)

        self.setWindowIcon(QtGui.QIcon("files\\ico\\order_works.ico"))

        self.main_widget = QtWidgets.QWidget()
        self.vertical_layout = QtWidgets.QVBoxLayout(self.main_widget)
        self.main_widget.setLayout(self.vertical_layout)
        self.setCentralWidget(self.main_widget)

        self.horizontal_layout = QtWidgets.QHBoxLayout()
        #self.horizontal_layout.setSpacing(30)

        spacerItem1 = QtWidgets.QSpacerItem(10, 50)
        self.horizontal_layout.addItem(spacerItem1)

        self.label_part1 = QtWidgets.QLabel(self.main_widget)
        '''font = QtGui.QFont()
        font.setFamily("Bahnschrift SemiLight SemiConde")
        font.setPointSize(12)'''
        self.label_part1.setFont(self.font)
        self.label_part1.setObjectName("label_part1")
        self.label_part1.setText("Укажите порядок выполнения работ.")
        self.horizontal_layout.addWidget(self.label_part1)

        spacerItem2 = QtWidgets.QSpacerItem(20, 50)
        self.horizontal_layout.addItem(spacerItem2)

        self.pushButton_building = QtWidgets.QPushButton(self.main_widget)
        self.pushButton_building.setMinimumSize(75, 27)
        self.pushButton_building.setObjectName("pushButton_building")
        self.pushButton_building.setText("Далее")
        self.pushButton_building.clicked.connect(
            lambda: self.check_buttons(num_str_in_table))
        self.pushButton_building.setFont(self.font)
        background_color = "#51d77a"
        border_color = "#2d8849"
        color = "#2d8849"
        background_color_hover = "#75e998"
        border_color_hover = "#2d8849"
        color_hover = "#2d8849"
        background_color_press = "#17c14b"
        border_color_press = "#2d8849"
        color_press = "#75e998"
        border_width = "1px"
        border_radius = "2px"
        button_style = StyleWidgets()
        button_style.properties_button(self.pushButton_building, background_color,
                                       border_color, color,
                                       background_color_hover, border_color_hover, color_hover,
                                       background_color_press, border_color_press, color_press,
                                       border_width, border_radius)
        self.horizontal_layout.addWidget(self.pushButton_building)

        spacerItem3 = QtWidgets.QSpacerItem(20, 50)
        self.horizontal_layout.addItem(spacerItem3)

        self.label_part2 = QtWidgets.QLabel(self.main_widget)
        font = QtGui.QFont()
        font.setFamily("Bahnschrift SemiLight SemiConde")
        font.setPointSize(13)
        self.label_part2.setFont(font)
        self.label_part2.setObjectName("label_part1")
        self.label_part2.setText("Этапы построения сетевого графика \nоткроются в браузере.")
        self.label_part2.setVisible(False)
        self.label_part2.setStyleSheet("""
        color: '#b51a1a';
        background: '#f6a8a8';
        padding: 0 10px""")
        self.horizontal_layout.addWidget(self.label_part2)

        self.horizontal_layout.addStretch(1)

        self.vertical_layout.addLayout(self.horizontal_layout)


        # BOTTOM
        scrollAreaBottom = QtWidgets.QScrollArea(self.main_widget)
        scrollAreaBottom.setWidgetResizable(True)
        scrollAreaBottom.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOn)
        scrollAreaBottom.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOn)
        self.vertical_layout.addWidget(scrollAreaBottom)


        # Виджет для размещения кнопок
        self.widget_for_buttons = QtWidgets.QWidget(scrollAreaBottom)
        self.widget_for_buttons.setFixedWidth(5900)
        self.widget_for_buttons.setFixedHeight(3200)
        # self.widget_for_task.setStyleSheet("""background: green;""")
        scrollAreaBottom.setWidget(self.widget_for_buttons)

        # Стиль шрифта для кнопок
        if small_font == False:
            self.font = QtGui.QFont()
            self.font.setFamily("Bahnschrift SemiLight SemiConde")
            self.font.setPointSize(13)
        else:
            self.font = QtGui.QFont()
            self.font.setFamily("Bahnschrift SemiLight SemiConde")
            self.font.setPointSize(10)

        global lst_btn, lst_btn_col, lst_btn_row
        lst_btn, lst_btn_col, lst_btn_row = [], [] ,[]

        for i in range(0, num_str_in_table - 1):
            for j in range(1, num_str_in_table):
                xJ = (j - 1)*59 + 40
                if i < j:
                    yI = 40 + i*31
                    #print(str(i + 1) + "⇾" + str(j + 1), xJ, yI)
                    globals()["btn" + str(i+1) + "_" + str(j+1)] = QtWidgets.QPushButton(self.widget_for_buttons)
                    globals()["btn" + str(i+1) + "_" + str(j+1)].setGeometry(QtCore.QRect(xJ, yI, 60, 32))
                    globals()["btn" + str(i+1) + "_" + str(j+1)].setFont(self.font)
                    globals()["btn" + str(i+1) + "_" + str(j+1)].setStyleSheet('''
                                                background-color: "#cccccc";
                                                border-style: solid;
                                                border-color: "#909090";
                                                border-width: 1px;
                                                border-radius: 0px;
                                                color: "#909090";''')
                    globals()["btn" + str(i+1) + "_" + str(j+1)].setText(str(i + 1) + "⇾" + str(j + 1))

        # Обработка нажатой кнопки
        for i in range(0, num_str_in_table - 1):
            for j in range(1, num_str_in_table):
                if i < j:
                    globals()["btn" + str(i+1) + "_" + str(j+1)].clicked.connect(self.pressed_button)

    # Метод обработки нажатой кнопки
    def pressed_button(self):
        global numBtn
        sender = self.sender() # устанавливаем какой виджет является отправителем сигнала
        textButton = sender.text()
        index = textButton.find("⇾") # находим индекс
        btn_text_one_num = textButton[:index]
        btn_text_two_num = textButton[index + 1:]
        #print(index, btn_text_one_num, btn_text_two_num)
        if not((btn_text_one_num + "_" + btn_text_two_num) in lst_btn):
            lst_btn.append(btn_text_one_num + "_" + btn_text_two_num)
            lst_btn_row.append(btn_text_one_num)
            lst_btn_col.append(btn_text_two_num)

            globals()["btn" + str(int(btn_text_one_num)) +\
                      "_" + str(int(btn_text_two_num))].setStyleSheet('''
            background-color: "#6cbef1";
            border-color: "#155b87";
            color: "#155b87";
            border-style: solid;
            border-width: 1px;
            border-radius: 0px;
            ''')
        else:
            lst_btn.remove(btn_text_one_num + "_" + btn_text_two_num)
            lst_btn_row.remove(btn_text_one_num)
            lst_btn_col.remove(btn_text_two_num)

            globals()["btn" + str(int(btn_text_one_num)) + \
                      "_" + str(int(btn_text_two_num))].setStyleSheet('''
            background-color: "#cccccc";
            border-color: "#909090";
            color: "#909090";
            border-style: solid;
            border-width: 1px;
            border-radius: 0px;''')

        #print("Содержимое списков:", lst_btn, lst_btn_row, lst_btn_col)

    def check_buttons(self, num_str_in_table):
        '''
        Метод проверяет, чтобы у каждой работы была предшествующая.
        '''
        if len(set(lst_btn_col)) == num_str_in_table - 1 and \
                len(set(lst_btn_row)) == num_str_in_table - 1:
            self.label_part2.setVisible(True)
            self.building.calculation_of_indicators(num_str_in_table)
        else:
            self.msg_box_win_order_of_work.setIcon(QtWidgets.QMessageBox.Warning)
            self.msg_box_win_order_of_work.setWindowTitle("Предупреждение")
            self.msg_box_win_order_of_work.setText("У каждой работы должна быть предшествующая!")
            self.msg_box_win_order_of_work.setFont(self.font)
            self.left = 670
            self.top = 250
            self.msg_box_win_order_of_work.setGeometry(self.left, self.top, 700, 550)
            self.msg_box_win_order_of_work.setWindowIcon(QtGui.QIcon("files\\ico\\order_works.ico"))
            self.msg_box_win_order_of_work.exec()

class BuildingNetworkGraph():
    def calculation_of_indicators(self, num_str_in_table):
        '''
        Вычисление показателей:
        раннее начало - РН (ES)
        ранне окончание - РО (EF)
        резервы - R
        позднее начало - ПН (LS)
        позднее окончание - ПО (LF)
        '''
        global dict_in_vtx, dict_vtx_value, dict_in_vtx_rev, dict_vtx_value_rev
        dict_in_vtx, dict_vtx_value, dict_in_vtx_rev, dict_vtx_value_rev = {}, {}, {}, {}
        # dict_in_vtx     - хранит вершины, которые входят в текущую при прямом проходе
        # dict_vtx_value  - хранит РН, T, РО, описание
        # dict_in_vtx_rev - хранит вершины, которые входят в текущую при обратном проходе
        # dict_vtx_value_rev - хранит ПН, R, ПО
        dict_vtx_value.update({1: [0, lst_work_day[0], lst_work_day[0],
                                   lst_work_description[0]]})

        #print("Расчет характеристик:")
        #print(lst_work_description, lst_work_day)
        global lstGet_dict_in_vtx, lst_btn_sort, lst_btn_sort2
        lstGet_dict_in_vtx, lst_btn_sort, lst_btn_sort2 = [], [], []
        # Сортировка по второй цифре
        lst_btn_sort = sorted(lst_btn, key=lambda last: last[2])
        #print("lst_btn_sort:", lst_btn_sort)

        # Нахождение вершин которые ведут к текущей
        for i in range(0, len(lst_btn_sort)):
            # Разделяем отсортированные наименования с кнопок
            lst_btn_sort2 = lst_btn_sort[i].split('_')
            #print('lst_btn_sort2:',lst_btn_sort2)
            # Считываем, что хранится в dictInVtx для того, чтобы к
            # этому содержимому добавить новый элемент
            lstGet_dict_in_vtx = dict_in_vtx.get(int(lst_btn_sort2[1]))
            #print("lstGet_dictInVtx:",lstGet_dictInVtx)
            #print("dict_in_vtx:", dict_in_vtx)
            if lstGet_dict_in_vtx != None:
                dict_in_vtx.update({int(lst_btn_sort2[1]): lstGet_dict_in_vtx + \
                                                           [int(lst_btn_sort2[0])]})
            else:
                dict_in_vtx.update({int(lst_btn_sort2[1]): [int(lst_btn_sort2[0])]})
        print("dict_in_vtx:",dict_in_vtx, '\n')

        # Заполнение значений РН, РО (прямой проход)
        for t in range(2, num_str_in_table + 1):
            # Если будет несколько путей в вершину,
            # то в этом списке будут храниться значения РО, входящие в вершину
            lst_some_ES = []
            # print(t,len(dictInVtx[t]))

            # Если в вершину входит два пути, то берем max из РО
            if len(dict_in_vtx[t]) > 1:
                for t2 in range(0, len(dict_in_vtx[t])):
                    ES = dict_vtx_value[dict_in_vtx[t][t2]][2]
                    lst_some_ES.append(ES)
                    dict_vtx_value.update({t: [max(lst_some_ES),
                                               lst_work_day[t - 1],
                                               max(lst_some_ES) + lst_work_day[t - 1],
                                               lst_work_description[t - 1]]})
            else:
                ES = dict_vtx_value[dict_in_vtx[t][0]][2]
                dict_vtx_value.update({t: [ES,
                                           lst_work_day[t - 1],
                                           ES + lst_work_day[t - 1],
                                           lst_work_description[t - 1]]})
        print("dict_vtx_value:", dict_vtx_value)

        # Нахождение вершин, пути от которых ведут к текущей (обратный проход)
        lstBtnSortRev = sorted(lst_btn)
        # Сортировка по первой цифре
        #print("lstBtnSortRev:", lstBtnSortRev)
        for l in range(0, len(lstBtnSortRev)):
            # Значение по ключу(номеру вершины)
            lst_btn_sort2_rev = lstBtnSortRev[l].split('_')
            #print("lstBtnSort2Rev:", lstBtnSort2Rev)
            lstGet_dict_in_vtx_rev = dict_in_vtx_rev.get(int(lst_btn_sort2_rev[0]))
            #print("lstGet_dictInVtxRev:", lstGet_dictInVtxRev)
            if lstGet_dict_in_vtx_rev != None:
                dict_in_vtx_rev.update({int(lst_btn_sort2_rev[0]): lstGet_dict_in_vtx_rev + \
                                                                [int(lst_btn_sort2_rev[1])]})
            else:
                dict_in_vtx_rev.update({int(lst_btn_sort2_rev[0]):
                                         [int(lst_btn_sort2_rev[1])]})
        print("dict_in_vtx_rev:", dict_in_vtx_rev, "\n")


        # Заполнение вершин (проход в обратную сторону)
        dict_vtx_value_rev.update({num_str_in_table: [dict_vtx_value[num_str_in_table][0], 0,
                                             dict_vtx_value[num_str_in_table][2]]})
        t = num_str_in_table - 1
        while t != 0:
            lst_some_LF = []
            # Если в вершину входит два пути
            if len(dict_in_vtx_rev[t]) > 1:
                for t2 in range(0, len(dict_in_vtx_rev[t])):
                    LF = dict_vtx_value_rev[dict_in_vtx_rev[t][t2]][0]
                    ES = dict_vtx_value[t][0]
                    T = dict_vtx_value[t][1]
                    lst_some_LF.append(LF)
                    LS = min(lst_some_LF) - T
                    dict_vtx_value_rev.update({t: [min(lst_some_LF) - T, LS - ES,
                                                min(lst_some_LF)]})
            else:
                LF = dict_vtx_value_rev[dict_in_vtx_rev[t][0]][0]
                ES = dict_vtx_value[t][0]
                T = dict_vtx_value[t][1]
                LS = LF - T
                dict_vtx_value_rev.update({t: [LF - T, LS - ES, LF]})
            t -= 1
        print("dict_vtx_value_rev:", dict_vtx_value_rev)
        self.network_graphViz(num_str_in_table)

    def network_graphViz(self, num_str_in_table):
        '''
        Построение сетевого графика, нахождение
        путей для вычисления резервов.
        Создание промежуточных сетевых графиков
        для добавления на HTML страницу.
        Формирование кода HTML страницы.
        '''
        # --------------------------------------------------------------
        os.environ["PATH"] += os.pathsep + "files\\Graphviz2.34\\bin\\"
        file = open("files\\html\\NetworkGraph.gv", "w", encoding="utf-8")
        file.write('''\
        digraph NetworkGraph
        {
           //graph [charset = "utf8"]
           rankdir = LR
           layout = dot
           splines = spline
           node [style = "filled, bold", fillcolor = "#f2f4f7", fontname = "Arial"]
           edge [penwidth = 2]
        ''')

        # Объединение значений прямого и обратного проходов
        global dict_vtx_value_all
        dict_vtx_value_all = {}
        for i in range(1, num_str_in_table + 1):
            dict_vtx_value_all.update({i: dict_vtx_value[i] + dict_vtx_value_rev[i]})
        #print("dict_vtx_value_all:", dict_vtx_value_all)

        lst_color = []
        for i in range(1, num_str_in_table + 1):
            lst_color.append(color_generate())

        # Перенос слов в вершинах
        for i in range(1, num_str_in_table + 1):
            if len(dict_vtx_value_all[i][3]) > 15:
                lst_str = dict_vtx_value_all[i][3].split(' ')
                str_join = lst_str[0]
                for t in range(1, len(lst_str)):
                    if len(lst_str[t - 1] + lst_str[t]) > 10:
                        str_join += "\\n " + lst_str[t] + " "
                    else:
                        str_join += " " + lst_str[t] + " "
                dict_vtx_value_all.update({i: [dict_vtx_value[i][0]] + \
                                           [dict_vtx_value[i][1]] + \
                                           [dict_vtx_value[i][2]] + \
                                           [str_join] + \
                                           dict_vtx_value_rev[i]})
        #print("После изменений (слова с переносом):", dict_vtx_value_all)

        # Запись вершин в файл
        for i in range(1, num_str_in_table + 1):
            file.write(u"   v" + str(i) + " [fontname = \"Arial\"," + \
                       "color = \"" + str(lst_color[i - 1]) + \
                       "\", shape = record, label = \"{ " + \
                       str(dict_vtx_value_all[i][0]) + " | " + \
                       str(dict_vtx_value_all[i][1]) + " | " + \
                       str(dict_vtx_value_all[i][2]) + " } | " + " <mI" + str(i) + "> " + \
                       str(i) + ". " + str(dict_vtx_value_all[i][3]) + " | {" + \
                       str(dict_vtx_value_all[i][4]) + " | " + \
                       str(dict_vtx_value_all[i][5]) + " | " + \
                       str(dict_vtx_value_all[i][6]) + " }\"]\n")

        # Список связей вершин
        lst_link_vtx = []
        for l in range(0, len(lst_btn_sort)):
            # Значение по ключу(номеру вершины)
            lst_link_vtx.append(lst_btn_sort[l].split('_'))
        print("lst_link_vtx:",lst_link_vtx)

        # Преобразование списка списков lst_link_vtx в кортеж
        tuple_lst_link_vtx = []
        for t in range(0, len(lst_btn_sort)):
            tuple_lst_link_vtx.append((int(lst_link_vtx[t][0]),
                                       int(lst_link_vtx[t][1])))
        print(tuple_lst_link_vtx)

        # Нахождение всех путей для вычисления резервов
        #G = nx.Graph()
        #G.add_edges_from(tuple_lst_link_vtx)

        #reserves = list(nx.all_simple_paths(G, source=1, target=num_str_in_table))
        #print("reserves:", reserves)
        G = nx.DiGraph(tuple_lst_link_vtx)
        lst_reserves_int = list(nx.all_simple_paths(G, source=1, target=num_str_in_table))
        print("reserves:", lst_reserves_int,"\nКоличество путей:", len(lst_reserves_int))

        # Запись вершин в файл
        for i in range(0, len(lst_btn_sort)):
            file.write("   v" + lst_link_vtx[i][0] + ":<mI" + str(lst_link_vtx[i][0]) +\
                       "> -> v" + lst_link_vtx[i][1] + ":<mI" + str(lst_link_vtx[i][1]) +\
                       "> [color = \"" + str(lst_color[int(lst_link_vtx[i][0]) - 1]) +\
                       "\"]\n")
        file.write("\n}")
        file.close()
        os.system(r"files\\startGV.bat files\\html\\NetworkGraph.gv")


        # Построение промежуточных сетевых графиков (схема с названием работ
        # без параметров)
        fileGV01 = open("files\\html\\NetworkGraph01.gv", "w", encoding="utf-8")
        fileGV01.write('''\
        digraph NetworkGraph01
        {
           //graph [charset = "utf8"]
           rankdir = LR
           layout = dot
           splines = spline
           node [style = "filled, bold", fillcolor = "#f2f4f7", fontname = "Arial"]
           edge [penwidth = 2]
        ''')
        # --------------------
        # Запись вершин в файл
        for i in range(1, num_str_in_table + 1):
            fileGV01.write(u"   v" + str(i) + " [fontname = \"Arial\"," + \
                           "color = \"" + str(lst_color[i - 1]) + \
                           "\", shape = record, label = \"{ " + \
                           "" + " | " + \
                           "" + " | " + \
                           "" + " } | " + " <mI" + str(i) + "> " + \
                           str(i) + ". " + str(dict_vtx_value_all[i][3]) + " | {" + \
                           "" + " | " + \
                           "" + " | " + \
                           "" + " }\"]\n")

        for i in range(0, len(lst_btn_sort)):
            fileGV01.write("   v" + lst_link_vtx[i][0] + ":<mI" + str(lst_link_vtx[i][0]) + \
                           "> -> v" + lst_link_vtx[i][1] + ":<mI" + str(lst_link_vtx[i][1]) + \
                           "> [color = \"" + str(lst_color[int(lst_link_vtx[i][0]) - 1]) + \
                           "\"]\n")
        fileGV01.write("\n}")
        fileGV01.close()
        os.system(r"files\\startGV.bat files\\html\\NetworkGraph01.gv")


        # Построение промежуточных сетевых графиков (схема с ранним началом и
        # ранним окончанием работ)
        fileGV02 = open("files\\html\\NetworkGraph02.gv", "w", encoding="utf-8")
        fileGV02.write('''\
        digraph NetworkGraph02
        {
           //graph [charset = "utf8"]
           rankdir = LR
           layout = dot
           splines = spline
           node [style = "filled, bold", fillcolor = "#f2f4f7", fontname = "Arial"]
           edge [penwidth = 2]
        ''')

        # Запись вершин в файл
        for i in range(1, num_str_in_table + 1):
            fileGV02.write(u"   v" + str(i) + " [fontname = \"Arial\"," + \
                           "color = \"" + str(lst_color[i - 1]) + \
                           "\", shape = record, label = \"{ " + \
                           str(dict_vtx_value_all[i][0]) + " | " + \
                           str(dict_vtx_value_all[i][1]) + " | " + \
                           str(dict_vtx_value_all[i][2]) + " } | " + \
                           " <mI" + str(i) + "> " + \
                           str(i) + ". " + str(dict_vtx_value_all[i][3]) + " | {" + \
                           "" + " | " + \
                           "" + " | " + \
                           "" + " }\"]\n")

        for i in range(0, len(lst_btn_sort)):
            fileGV02.write("   v" + lst_link_vtx[i][0] + ":<mI" + str(lst_link_vtx[i][0]) + \
                           "> -> v" + lst_link_vtx[i][1] + ":<mI" + str(lst_link_vtx[i][1]) + \
                           "> [color = \"" + str(lst_color[int(lst_link_vtx[i][0]) - 1]) + \
                           "\"]\n")
        fileGV02.write("\n}")
        fileGV02.close()
        os.system(r"files\\startGV.bat files\\html\\NetworkGraph02.gv")

        # Вычисление резервов
        # lst_str_reserves - список всех резервов
        # lst_str_reserve - список, который входит в список всех резервов (1 резерв)
        lst_str_reserves = []
        lst_int_reserves = []
        for i in range(0, len(lst_reserves_int)):
            lst_str_reserve = []
            lst_int_reserve = []
            for l in range(0, len(lst_reserves_int[i])):
                k = lst_reserves_int[i][l]
                lst_str_reserve.append(str(dict_vtx_value_all[k][1]))
                lst_int_reserve.append(dict_vtx_value_all[k][1])
            lst_str_reserves.append(lst_str_reserve)
            lst_int_reserves.append(lst_int_reserve)

        # Запись в файл html
        if num_str_in_table < 7:
            width_img = "70%"
        elif num_str_in_table == 7:
            width_img = "70%"
        else:
            width_img = "95%"

        fileHTML = open("files\\html\\Info.html", "w", encoding="utf-8")
        fileHTML.write('''<!DOCTYPE html>
        <html lang="ru">
          <head>
            <meta charset="utf-8">
            <title>Сетевой график</title>
        	<link rel = "stylesheet" href = "Output.css">
        	<link rel = "shorctcut icon" href = "NetDiag.ico">
          </head>
          <body>
        	<div class = "text-content">
        	 <h1 class = "align-center">Построение сетевого графика</h1>
        	 <p><b>Сетевой график</b> – это ориентированный граф, где в вершинах 
        	 располагаются выполняемые работы, дугами – изображается связь между ними.  
        	 Каждая вершина содержит несколько параметров, используя необходимые 
        	 формулы можно рассчитать значения в каждой ячейке (рис. 1, формулы (1-5)).</p>
        	 <br />
        	 <table cellspacing = "0" cellpadding = "3" width = "50%" align = "center" class = "imgTable">
        	   <tr align = "center">
        	     <td><img src="https://math.now.sh?from=T%5E%7B%D0%A0%D0%9D%7D_i" /></td>
        		 <td><img src="https://math.now.sh?from=T_i" /></td>
        		 <td><img src="https://math.now.sh?from=T%5E%7B%D0%A0%D0%9E%7D_i" /></td>
        	   </tr> 
        	   <tr align = "center">
        	     <td colspan = "3"><Номер работы>.<Название задачи></td>
        	   </tr>
        	   <tr align = "center">
        	     <td><img src="https://math.now.sh?from=T%5E%7B%D0%9F%D0%9D%7D_i" /></td>
        		 <td><img src="https://math.now.sh?from=R_i" /></td>
        		 <td><img src="https://math.now.sh?from=T%5E%7B%D0%9F%D0%9E%7D_i" /></td>
        	   </tr>
        	 </table>
        	 <p class = "smallText"> Рис. 1. Вершина сетевого графика с параметрами </p>
        	 <!-- Формулы под картинкой с параметрами -->
        	 <p class = "align-center">
        	   <img src="https://math.now.sh?from=T%5E%7B%D0%A0%D0%9D%7D_i%20%3D%20max%28T%5E%7B%D0%A0%D0%9E%7D_%7Bk%7D%29" />
        	   <img src="https://math.now.sh?from=%281%29" style = "padding: 0 3px 3px 10px" />
        	   <img style = "padding: 0 0 0 80px"
        	   src="https://math.now.sh?from=T%5E%7B%D0%9F%D0%9D%7D_i%20%3D%20T%5E%7B%D0%9F%D0%9E%7D_i%20-%20T_i" />
        	   <img src="https://math.now.sh?from=%283%29" style = "padding: 0 3px 3px 20px" />
        	   <br />
        	   <img style = "padding: 0 0 0 20px"
        	   src="https://math.now.sh?from=T%5E%7B%D0%A0%D0%9E%7D_i%20%3D%20T%5E%7B%D0%A0%D0%9D%7D_i%20%2B%20T_i" />
        	   <img src="https://math.now.sh?from=%282%29" style = "padding: 0 3px 3px 20px" />
        	   <img style = "padding: 0 0 0 80px"
        	   src="https://math.now.sh?from=T%5E%7B%D0%9F%D0%9E%7D_i%20%3D%20min%28T%5E%7B%D0%9F%D0%9D%7D_%7Bj%7D%29" />
        	   <img src="https://math.now.sh?from=%284%29" style = "padding: 0 3px 3px 10px" />
        	   <br />
        	   <img 
        	   src="https://math.now.sh?from=R_i%20%3D%20T%5E%7B%D0%9F%D0%9E%7D_i-T%5E%7B%D0%A0%D0%9E%7D_i%20%3D%20T%5E%7B%D0%9F%D0%9D%7D_i-T%5E%7B%D0%A0%D0%9D%7D_i%20" />
        	   <img src="https://math.now.sh?from=%285%29" style = "padding: 0 3px 3px 20px" />
        	   <p> где <img style = "padding: 20px 17px 0 8px" src="https://math.now.sh?from=i" /> 
        	           - текущая работа <br />
        	           <img  class = "textFormulas" src="https://math.now.sh?from=k%5Cin%20%5Cleft%5C%7Bi-1%2Ci-2%2C...%2Ci-N%5Cright%5C%7D" />
        	           <br /><img class = "textFormulas" src="https://math.now.sh?from=j%5Cin%20%5Cleft%5C%7Bi%2B1%2Ci%2B2%2C...%2Ci%2BN%5Cright%5C%7D" />
        	           <br />
        	           <img class = "textFormulas" 
        			        src="https://math.now.sh?from=T%5E%7B%D0%A0%D0%9D%7D_i" />
        			   - раннее начало		
        			   <br />
        			   <img style = "padding: 0px 17px 0 50px"
        			        src="https://math.now.sh?from=T_i" />
        			   - длительность работы
        			   <br />
        			   <img class = "textFormulas" 
        			        src="https://math.now.sh?from=T%5E%7B%D0%A0%D0%9E%7D_i" />
        			   - раннее окончание		
        			   <br />
        			   <img class = "textFormulas" 
        			        src="https://math.now.sh?from=T%5E%7B%D0%9F%D0%9D%7D_i" />
        			   - позднее начало
        			   <br />
        			   <img class = "textFormulas" 
        			        src="https://math.now.sh?from=T%5E%7B%D0%9F%D0%9E%7D_i" />
        		       - позднее окончание
        			   <br />
        			   <img style = "padding: 0px 17px 0 50px"
        			        src="https://math.now.sh?from=R_i" />
        			   - временной резерв
        	   </p>
        	 </p>
             <p> Описание и длительность необходимых работ представлены в таблице 1.</p>
             <p class = "smallText"> Таблица 1. Описание выполняемых работ </p>
             <div class="TableScroll">
        	    <table cellspacing = "0" cellpadding = "3" width = "50%" align = "center">
        	      <tbody>
        	        <tr align = "center" style = "background-color: #3aa9fe; ">
        	          <th>Номер <br /> работы </th>
        		      <th>Описание работы </th>
        		      <th>Длительность <br />(в днях) </th>
        	        </tr>
        ''')
        # Запись таблицы в файл html
        for i in range(0, num_str_in_table):
            fileHTML.write('''	        <tr class = "selectionColor">
        	          <td align = "center">''' + str(i + 1) + '''</td>
        		      <td>''' + lst_work_description[i] + '''</td>
        		      <td align = "center">''' + str(lst_work_day[i]) + '''</td>
        	        </tr>''')
        fileHTML.write('''		   </tbody>
        	     </table>
        	  </div>
        	  <p> Последовательность выполнения работ, которая была указана в окне
        	  ввода программы представлена на схеме (рис. 2). </p>
        	  <a href = "NetworkGraph01.svg"><img src = "NetworkGraph01.jpg"
        	     width = ''' + width_img + '''
        	  title = "Открыть в полном размере" class = "imgSetting"/></a>
        	  <p class = "smallText"> Рис. 2. Последовательность необходимых работ </p>
        	  <p> Рассчитаем значения параметров <b>раннего начала</b>
        	  и <b>раннего окончания</b> работ, используя формулы (1), (2). В первой
        	  вершине раннее начало работы будет равно нулю. Все остальные вычисления
        	  размещены в блоке ниже и на сетевом графике (рис. 3):</p>
        	  <!-- формулы в div с полосой прокрутки -->
        	  <div class = "WinFormulas">
        	    <p style = "padding: 0">
        	      <img src="https://math.now.sh?from=T%5E%7B%D0%A0%D0%9D%7D_1%20%3D%200" />
        	    </p>
        	    <p style = "padding: 0 0 20px 0">
        	      <img src="https://math.now.sh?from=
        		  T%5E%7B%D0%A0%D0%9E%7D_1%20%3D
        		  %20T%5E%7B%D0%A0%D0%9D%7D_1%20%2B%20T_1
        		  %3D%20''' + str(dict_vtx_value_all[1][0]) + \
                       '''%20%2B%20''' + str(dict_vtx_value_all[1][1]) + \
                       '''%20%3D%20''' + \
                       str(dict_vtx_value_all[1][0] + dict_vtx_value_all[1][1]) + '''" />
        	    </p>''')
        for i in range(2, num_str_in_table + 1):
            fileHTML.write('''
        		<p style = "padding: 0">''' + \
                           maxValue_HTML(i) + \
                           '''
                         </p>
                         <p style = "padding: 0 0 20px 0">
                           <img src="https://math.now.sh?from=
                           T%5E%7B%D0%A0%D0%9E%7D_%7B''' + str(i) + \
                           '''%7D%20%3D%20T%5E%7B%D0%A0%D0%9D%7D_%7B''' + str(i) + \
                           '''%7D%20%2B%20T_%7B''' + str(i) + \
                           '''%7D%3D%20''' + str(dict_vtx_value_all[i][0]) + \
                           '''%20%2B%20''' + str(dict_vtx_value_all[i][1]) + \
                           '''%20%3D%20''' + \
                           str(dict_vtx_value_all[i][0] + dict_vtx_value_all[i][1]) + \
                           '''" />
        	    </p>''')
        fileHTML.write('''
        	  </div>
        	  <a href = "NetworkGraph02.svg"><img src = "NetworkGraph02.jpg"
        	     width = "''' + width_img + '''"
        	  title = "Открыть в полном размере" class = "imgSetting"/></a>
        	  <p class = "smallText"> Рис. 3. Результаты вычислений
        	  раннего начала и окончания работ</p>
        	  <p> Следующий этап построения сетевого графика заключается в
        	  нахождении <b>позднего окончания</b>, <b>позднего начала</b> и
        	  <b>резерва времени</b> проводимых работ. В конечной работе <b>«''' + \
                       str(num_str_in_table) + ". " + \
                       lst_work_description[num_str_in_table - 1] + \
                       '''»</b> значения позднего 
                   начала и окончания работ соответствуют раннему началу и раннему
                   окончанию работ. Все расчеты (рис. 4) проводятся от последней работы в
                   обратном направлении стрелочек по формулам (3), (4), (5):</p>
                   <div class = "WinFormulas">''')
        fileHTML.write('''
        	    <p style = "padding: 0">
        	      <img src="https://math.now.sh?from=
        	      T%5E%7B%D0%9F%D0%9E%7D_%7B''' + str(num_str_in_table) + \
                       '''%7D%20%3D%20''' + str(dict_vtx_value_all[num_str_in_table][6]) + '''" />
                   </p>''' + \
                       '''
                           <p style = "padding: 0">
                         <img src="https://math.now.sh?from=
                         T%5E%7B%D0%9F%D0%9D%7D_%7B''' + str(num_str_in_table) + \
                       '''%7D%20%3D%20''' + str(dict_vtx_value_all[num_str_in_table][4]) + '''" />
                    </p>
        	    <p style = "padding: 0 0 20px 0">
        	      <img src="https://math.now.sh?from=
        	      R_%7B''' + str(i) + '''%7D%20%3D%20T%5E%7B%D0%9F%D0%9D%7D_%7B''' + \
                       str(num_str_in_table) + '''%7D-T%5E%7B%D0%A0%D0%9D%7D_%7B''' + \
                       str(num_str_in_table) + \
                       '''%7D%20%3D%20''' + str(dict_vtx_value_all[num_str_in_table][4]) + \
                       '''%20-%20''' + str(dict_vtx_value_all[num_str_in_table][0]) + \
                       '''%20%3D%20''' + \
                       str(dict_vtx_value_all[num_str_in_table][5]) + '''" />
        	    </p>''')
        i = num_str_in_table - 1
        while i != 0:
            fileHTML.write('''
        		<p style = "padding: 0">''' + \
                           minValue_HTML(i) + \
                           '''</p>
                           <p style = "padding: 0">
                             <img src="https://math.now.sh?from=
                             T%5E%7B%D0%9F%D0%9D%7D_%7B''' + str(i) + \
                           '''%7D%20%3D%20T%5E%7B%D0%9F%D0%9E%7D_%7B''' + \
                           str(i) + '''%7D-T_%7B''' + str(i) + \
                           '''%7D%20%3D%20''' + str(dict_vtx_value_all[i][6]) + \
                           '''%20-%20''' + str(dict_vtx_value_all[i][1]) + '''%20%3D%20''' + \
                           str(dict_vtx_value_all[i][4]) + '''" />
        	    </p>
        	    <p style = "padding: 0 0 20px 0">
        	      <img src="https://math.now.sh?from=
        	      R_%7B''' + str(i) + '''%7D%20%3D%20T%5E%7B%D0%9F%D0%9D%7D_%7B''' + \
                           str(i) + '''%7D-T%5E%7B%D0%A0%D0%9D%7D_%7B''' + str(i) + \
                           '''%7D%20%3D%20''' + str(dict_vtx_value_all[i][4]) + \
                           '''%20-%20''' + str(dict_vtx_value_all[i][0]) + '''%20%3D%20''' + \
                           str(dict_vtx_value_all[i][5]) + '''" />
        	    </p>''')
            i -= 1
        fileHTML.write('''	
        	  </div>
        	  <a href = "NetworkGraph.svg"><img src = "NetworkGraph.jpg"
        	     width =  "''' + width_img + '''" title = "Открыть в полном размере"
                     class = "imgSetting"/></a>
        	  <p class = "smallText"> Рис. 4. Результаты построения
        	  сетевого графика</p>
        	  <p>Вычислим резервы времени по каждому из путей:</p>
        	  <div class = "WinFormulas">''')
        # -------------------------------------
        # Критический путь, резервы времени
        # LF_last_vtx - значение ПО (LF) из последней вершины
        LF_last_vtx = str(dict_vtx_value_all[num_str_in_table][6])
        plus = " + "
        arrow = "➜"

        # Преобразование путей в str
        # lst_reserves_str - все пути представлены в str в списках
        # lst_reserve_str - один из путей, входит в состав lst_reserves_str
        # --------------------------
        # Пример:
        # Преобразование reserves: [[1, 2, 3, 4, 5], [1, 2, 4, 5]]
        # в lst_reserves_str: [['1', '2', '3', '4', '5'], ['1', '2', '4', '5']]
        # ---------------------------------------------------------------------
        lst_reserves_str = []
        for i in range(0, len(lst_reserves_int)):
            lst_reserve_str = []
            for l in range(0, len(lst_reserves_int[i])):
                lst_reserve_str.append(str(lst_reserves_int[i][l]))
            lst_reserves_str.append(lst_reserve_str)

        # sum_str_lst_reserves_str - выполняет сложение строк из списка lst_str_reserves
        # для отображения на экране
        # sum_value_lst_reserves_int - сложение значений путей в спсике lst_reserves_int
        for i in range(0, len(lst_str_reserves)):
            sum_str_lst_reserves_str = plus.join(lst_str_reserves[i])
            sum_value_lst_reserves_int = sum(lst_int_reserves[i])
            reserves_with_arrows = arrow.join(lst_reserves_str[i])
            # ---------------------------------------
            str1 = '''
        	    <div class = "divIndent"> Путь ''' + \
                   '''<div class = "colorPath">''' + \
                   str(reserves_with_arrows) + "</div> : <br> <nobr> &nbsp &nbsp &nbsp &nbsp &nbsp"
            str_space = ""
            for space in range(0, len(str1)):
                str_space += " "
            # ---------------------------------------------------
            str2 = str_space + LF_last_vtx + " - " + "(" + sum_str_lst_reserves_str + ") = " + \
                   LF_last_vtx + " - " + str(sum_value_lst_reserves_int) + " = " + \
                   str(int(LF_last_vtx) - sum_value_lst_reserves_int) + '''</nobr></div>
        		<hr> '''
            fileHTML.write(str1 + str2)

            if int(LF_last_vtx) - sum_value_lst_reserves_int == 0:
                critical_path = i
                """fileHTML.write('''<div class = "divIndent">Критический путь: <div class = "colorCrPath">''' + \
                               str(arrow.join(lst_reserves_str[critical_path])) + "</div></div>")
                fileHTML.write("<hr>")"""

        fileHTML.write('''
                	  </div>''')
        fileHTML.write('''<div class = "divIndent_CrPath">Критический путь: <div class = "colorCrPath">''' + \
                       str(arrow.join(lst_reserves_str[critical_path])) + "</div></div>")
        fileHTML.write('''
           </div>
         </body>
        </html>
        ''')
        fileHTML.close()

        wb.open("files\\html\\Info.html", new=2)
        #os.system(r"files\\Info.html")

# Нахождение max значений раннего начала
def maxValue_HTML(i):
    if len(dict_in_vtx[i]) > 1:
        htmlReturn = '''
    	      <img src="https://math.now.sh?from=
                              T%5E%7B%D0%A0%D0%9D%7D_%7B''' + str(i) + \
                     '''%7D%20%3D%20max%28'''
        # max(T^{РО}_{i-1})
        for t in range(0, len(dict_in_vtx[i])):
            htmlReturn += '''T%5E%7B%D0%A0%D0%9E%7D_%7B''' + \
                          str(dict_in_vtx[i][t])
            if t + 1 < len(dict_in_vtx[i]):
                htmlReturn += '''%7D%2C'''
        # max(T^{РО}_{i-1}) = ...
        htmlReturn += '''%7D%29%20%3D%20max%28'''
        for t in range(0, len(dict_in_vtx[i])):
            htmlReturn += str(dict_vtx_value_all[dict_in_vtx[i][t]][2])
            if t + 1 < len(dict_in_vtx[i]):
                htmlReturn += '''%2C'''
        htmlReturn += '''%29%20%3D%20''' + \
                      str(dict_vtx_value_all[i][0]) + '''" />'''
    else:
        htmlReturn = '''
    	      <img src="https://math.now.sh?from=
                              T%5E%7B%D0%A0%D0%9D%7D_%7B''' + str(i) + \
                     '''%7D%20%3D%20''' + \
                     str(dict_vtx_value_all[i][0]) + '''" />'''
    return htmlReturn

# Нахождение min значений позднего начала
def minValue_HTML(i):
    if len(dict_in_vtx_rev[i]) > 1:
        htmlReturn2 = '''
    	      <img src="https://math.now.sh?from=
                              T%5E%7B%D0%9F%D0%9E%7D_%7B''' + str(i) + \
                      '''%7D%20%3D%20min%28'''
        # max(T^{РО}_{i-1})
        for t in range(0, len(dict_in_vtx_rev[i])):
            htmlReturn2 += '''T%5E%7B%D0%9F%D0%9D%7D_%7B''' + \
                           str(dict_in_vtx_rev[i][t])
            if t + 1 < len(dict_in_vtx_rev[i]):
                htmlReturn2 += '''%7D%2C'''
        # max(T^{РО}_{i-1}) = ...
        htmlReturn2 += '''%7D%29%20%3D%20min%28'''
        for t in range(0, len(dict_in_vtx_rev[i])):
            # Ищем в dictInVtxRev i-ую вершину со значением
            htmlReturn2 += str(dict_vtx_value_all[dict_in_vtx_rev[i][t]][4])
            if t + 1 < len(dict_in_vtx_rev[i]):
                htmlReturn2 += '''%2C'''
        htmlReturn2 += '''%29%20%3D%20''' + \
                       str(dict_vtx_value_all[i][6]) + '''" />'''
    else:
        htmlReturn2 = '''
    	      <img src="https://math.now.sh?from=
                              T%5E%7B%D0%9F%D0%9E%7D_%7B''' + str(i) + \
                      '''%7D%20%3D%20''' + \
                      str(dict_vtx_value_all[i][6]) + '''" />'''
    return htmlReturn2

#=====================================================================
'''def log_uncaught_exceptions(ex_cls, ex, tb):
    text = '{}: {}:\n'.format(ex_cls.__name__, ex)
    import traceback
    text += ''.join(traceback.format_tb(tb))
    print(text)
    QtWidgets.QMessageBox.critical(None, 'Error', text)
    quit()
import sys
sys.excepthook = log_uncaught_exceptions

app = QtWidgets.QApplication(sys.argv)'''
#=====================================================================

if __name__ == '__main__':
    import sys
    app = QtWidgets.QApplication(sys.argv)
    w = Ui_MainWindow()
    w.show()
    sys.exit(app.exec_())

